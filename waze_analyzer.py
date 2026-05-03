#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║  Waze Route Analyzer — ניתוח מסלולים אוטומטי               ║
║  ─────────────────────────────────────────────────────────── ║
║  שימוש:                                                     ║
║    python waze_analyzer.py <input.json> [output.xlsx]        ║
║                                                              ║
║  דוגמה:                                                      ║
║    python waze_analyzer.py waze_data.json                    ║
║    python waze_analyzer.py waze_data.json report_03_05.xlsx  ║
╚══════════════════════════════════════════════════════════════╝
"""

import json, sys, os, statistics
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ┌──────────────────────────────────────────────────────────┐
# │  הגדרת מסלולים — ניתן להוסיף/לערוך כאן                  │
# │  כל מסלול מוגדר עם:                                       │
# │    section  = אזור (דרום/מרכז/צפון)                       │
# │    name     = שם הכביש                                     │
# │    from/to  = מוצא/יעד                                     │
# │    distance_km  = מרחק בק"מ                                │
# │    free_flow_min = זמן חופשי בדקות                         │
# │    streets_dir1/2 = שמות רחובות ב-Waze לכל כיוון          │
# │    dir1/2_label = תווית כיוון                              │
# └──────────────────────────────────────────────────────────┘
ROUTES = [
    # ═══ דרום ═══
    {"section":"דרום","name":"כביש 10","from":"ניצנה","to":"אילת","distance_km":200,"free_flow_min":140,
     "streets_dir1":["10"],"streets_dir2":[],"dir1_label":"שני הכיוונים","dir2_label":""},
    {"section":"דרום","name":"כביש 25","from":"באר שבע","to":"שדרות","distance_km":35,"free_flow_min":30,
     "streets_dir1":["25 מזרח"],"streets_dir2":["25 מערב"],"dir1_label":"שדרות → באר שבע (מזרח)","dir2_label":"באר שבע → שדרות (מערב)"},
    {"section":"דרום","name":"כביש 40","from":"באר שבע","to":"תל אביב","distance_km":110,"free_flow_min":70,
     "streets_dir1":["40 צפון"],"streets_dir2":["40 דרום"],"dir1_label":"באר שבע → ת\"א (צפון)","dir2_label":"ת\"א → באר שבע (דרום)"},

    # ═══ מרכז ═══
    {"section":"מרכז","name":"כביש 1","from":"תל אביב","to":"ירושלים","distance_km":62,"free_flow_min":45,
     "streets_dir1":["1 מזרח"],"streets_dir2":["1 מערב"],"dir1_label":"תל אביב → ירושלים","dir2_label":"ירושלים → תל אביב"},
    {"section":"מרכז","name":"כביש 4","from":"אשדוד","to":"חיפה","distance_km":120,"free_flow_min":80,
     "streets_dir1":["4 צפון"],"streets_dir2":["4 דרום"],"dir1_label":"אשדוד → חיפה (צפון)","dir2_label":"חיפה → אשדוד (דרום)"},
    {"section":"מרכז","name":"כביש 5","from":"ת\"א","to":"אריאל","distance_km":40,"free_flow_min":30,
     "streets_dir1":["5 מזרח"],"streets_dir2":["5 מערב"],"dir1_label":"ת\"א → אריאל (מזרח)","dir2_label":"אריאל → ת\"א (מערב)"},
    {"section":"מרכז","name":"כביש 6","from":"קריית גת","to":"חדרה","distance_km":130,"free_flow_min":65,
     "streets_dir1":["6 צפון"],"streets_dir2":["6 דרום"],"dir1_label":"ק. גת → חדרה (צפון)","dir2_label":"חדרה → ק. גת (דרום)"},
    {"section":"מרכז","name":"כביש 41","from":"אשדוד","to":"ראשל\"צ","distance_km":20,"free_flow_min":18,
     "streets_dir1":["41 מזרח"],"streets_dir2":["41 מערב"],"dir1_label":"אשדוד → ראשל\"צ (מזרח)","dir2_label":"ראשל\"צ → אשדוד (מערב)"},
    {"section":"מרכז","name":"כביש 44","from":"אשדוד","to":"מודיעין","distance_km":40,"free_flow_min":35,
     "streets_dir1":["44 צפון"],"streets_dir2":["44 דרום"],"dir1_label":"אשדוד → מודיעין (צפון)","dir2_label":"מודיעין → אשדוד (דרום)"},
    {"section":"מרכז","name":"כביש 60","from":"באר שבע","to":"נצרת","distance_km":200,"free_flow_min":180,
     "streets_dir1":["60"],"streets_dir2":[],"dir1_label":"שני הכיוונים (מעורב)","dir2_label":""},
    {"section":"מרכז","name":"כביש 444","from":"ראש העין","to":"נחשונים","distance_km":15,"free_flow_min":12,
     "streets_dir1":["444 צפון"],"streets_dir2":["444 דרום"],"dir1_label":"נחשונים → ר\"ע (צפון)","dir2_label":"ר\"ע → נחשונים (דרום)"},
    {"section":"מרכז","name":"כביש 461","from":"אור יהודה","to":"יהוד","distance_km":8,"free_flow_min":8,
     "streets_dir1":["461 מזרח"],"streets_dir2":["461 מערב"],"dir1_label":"אור יהודה → יהוד (מזרח)","dir2_label":"יהוד → אור יהודה (מערב)"},

    # ═══ צפון ═══
    {"section":"צפון","name":"כביש 22","from":"חיפה מפרץ","to":"חיפה כרמל","distance_km":12,"free_flow_min":12,
     "streets_dir1":["22 צפון"],"streets_dir2":["22 דרום"],"dir1_label":"דרום → צפון","dir2_label":"צפון → דרום"},
    {"section":"צפון","name":"כביש 57","from":"נתניה","to":"טול כרם","distance_km":22,"free_flow_min":20,
     "streets_dir1":["57 מזרח"],"streets_dir2":["57 מערב"],"dir1_label":"נתניה → מזרח","dir2_label":"מזרח → נתניה"},
    {"section":"צפון","name":"כביש 65","from":"חדרה","to":"עפולה","distance_km":55,"free_flow_min":45,
     "streets_dir1":["65 מזרח"],"streets_dir2":["65 מערב"],"dir1_label":"חדרה → עפולה (מזרח)","dir2_label":"עפולה → חדרה (מערב)"},
    {"section":"צפון","name":"כביש 66","from":"מגידו","to":"עפולה","distance_km":20,"free_flow_min":18,
     "streets_dir1":["66 צפון"],"streets_dir2":["66 דרום"],"dir1_label":"מגידו → עפולה (צפון)","dir2_label":"עפולה → מגידו (דרום)"},
    {"section":"צפון","name":"כביש 70","from":"חיפה","to":"עכו","distance_km":25,"free_flow_min":20,
     "streets_dir1":["70 צפון"],"streets_dir2":["70 דרום"],"dir1_label":"חיפה → עכו (צפון)","dir2_label":"עכו → חיפה (דרום)"},
    {"section":"צפון","name":"כביש 75","from":"חיפה","to":"נצרת","distance_km":30,"free_flow_min":30,
     "streets_dir1":["75"],"streets_dir2":[],"dir1_label":"שני הכיוונים","dir2_label":""},
    {"section":"צפון","name":"כביש 89","from":"צפת","to":"נהריה","distance_km":42,"free_flow_min":40,
     "streets_dir1":["89"],"streets_dir2":[],"dir1_label":"שני הכיוונים (מעורב)","dir2_label":""},
    {"section":"צפון","name":"כביש 90","from":"אילת","to":"מטולה","distance_km":480,"free_flow_min":360,
     "streets_dir1":["90"],"streets_dir2":[],"dir1_label":"שני הכיוונים (מעורב)","dir2_label":""},
    {"section":"צפון","name":"כביש 98","from":"תל קציר","to":"מבוא חמה (גולן)","distance_km":25,"free_flow_min":25,
     "streets_dir1":["98"],"streets_dir2":[],"dir1_label":"שני הכיוונים","dir2_label":""},
    {"section":"צפון","name":"כביש 781","from":"כפר ביאליק","to":"קריות","distance_km":5,"free_flow_min":6,
     "streets_dir1":["781"],"streets_dir2":[],"dir1_label":"שני הכיוונים","dir2_label":""},
    {"section":"צפון","name":"כביש 804","from":"ראס אל-עין","to":"פקיעין","distance_km":12,"free_flow_min":15,
     "streets_dir1":["804"],"streets_dir2":[],"dir1_label":"שני הכיוונים","dir2_label":""},
    {"section":"צפון","name":"כביש 807","from":"מגדל","to":"רביד (כינרת)","distance_km":10,"free_flow_min":12,
     "streets_dir1":["807"],"streets_dir2":[],"dir1_label":"שני הכיוונים","dir2_label":""},
    {"section":"צפון","name":"כביש 866","from":"כרמיאל","to":"צפת","distance_km":30,"free_flow_min":35,
     "streets_dir1":["866"],"streets_dir2":[],"dir1_label":"שני הכיוונים (מעורב)","dir2_label":""},
    {"section":"צפון","name":"כביש 899","from":"יפתח","to":"גבול לבנון","distance_km":8,"free_flow_min":10,
     "streets_dir1":["899"],"streets_dir2":[],"dir1_label":"שני הכיוונים","dir2_label":""},
    {"section":"צפון","name":"כביש 8655","from":"פקיעין","to":"פקיעין חדשה","distance_km":5,"free_flow_min":6,
     "streets_dir1":["8655"],"streets_dir2":[],"dir1_label":"שני הכיוונים","dir2_label":""},
    {"section":"צפון","name":"כביש 8697","from":"רמות (גולן)","to":"רמות","distance_km":5,"free_flow_min":5,
     "streets_dir1":["8697"],"streets_dir2":[],"dir1_label":"שני הכיוונים","dir2_label":""},
]

# ┌──────────────────────────────────────────────────────────┐
# │  סטיילינג                                                │
# └──────────────────────────────────────────────────────────┘
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HDR_FILL = PatternFill('solid', fgColor='1F3864')
DATA_FONT = Font(name='Arial', size=10)
BOLD = Font(name='Arial', bold=True, size=10)
RED = Font(name='Arial', bold=True, color='C00000', size=10)
GREEN = Font(name='Arial', bold=True, color='006100', size=10)
ORANGE = Font(name='Arial', bold=True, color='BF6000', size=10)
BROWN = Font(name='Arial', bold=True, color='996600', size=10)
ALT = PatternFill('solid', fgColor='F2F7FB')
RED_BG = PatternFill('solid', fgColor='FCE4EC')
GREEN_BG = PatternFill('solid', fgColor='E8F5E9')
YELLOW_BG = PatternFill('solid', fgColor='FFF8E1')
ORANGE_BG = PatternFill('solid', fgColor='FFF3E0')
SEC_BG = PatternFill('solid', fgColor='D6E4F0')
BORDER = Border(left=Side(style='thin', color='C0C0C0'), right=Side(style='thin', color='C0C0C0'),
                top=Side(style='thin', color='C0C0C0'), bottom=Side(style='thin', color='C0C0C0'))

TIME_BINS = ['00:00-04:00','04:00-08:00','08:00-12:00','12:00-16:00','16:00-20:00','20:00-24:00']

def tbin(h): b=(h//4)*4; return f"{b:02d}:00-{b+4:02d}:00"
def hdr(ws, cols, row=1):
    for c, col in enumerate(cols,1):
        cl=ws.cell(row=row,column=c,value=col); cl.font=HDR_FONT; cl.fill=HDR_FILL
        cl.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cl.border=BORDER
def srow(ws, rn, nc, alt=False):
    for c in range(1,nc+1):
        cl=ws.cell(row=rn,column=c); cl.font=DATA_FONT; cl.border=BORDER
        cl.alignment=Alignment(vertical='center',horizontal='center')
        if alt: cl.fill=ALT
def sec_row(ws, rn, nc, text):
    ws.merge_cells(start_row=rn,start_column=1,end_row=rn,end_column=nc)
    c=ws.cell(row=rn,column=1,value=text)
    c.font=Font(name='Arial',bold=True,size=11,color='1F3864'); c.fill=SEC_BG
    c.alignment=Alignment(horizontal='right')
    for cc in range(1,nc+1): ws.cell(row=rn,column=cc).border=BORDER
def autow(ws, nc, mr=80):
    for c in range(1,nc+1):
        mx=0
        for r in range(1,min(ws.max_row+1,mr+2)):
            v=ws.cell(row=r,column=c).value
            if v is not None: mx=max(mx,len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width=min(mx+4,32)
def color_est(cell, est, ff):
    if not isinstance(est,(int,float)) or ff<=0: return
    ratio=est/ff
    if ratio>1.5: cell.fill=RED_BG; cell.font=RED
    elif ratio>1.2: cell.fill=ORANGE_BG; cell.font=ORANGE
    elif ratio>1.05: cell.fill=YELLOW_BG; cell.font=BROWN
    else: cell.fill=GREEN_BG; cell.font=GREEN
def color_status(cell, st):
    m={'חריג מאוד':(RED_BG,RED),'עמוס':(ORANGE_BG,ORANGE),'מתון':(YELLOW_BG,BROWN),'תקין':(GREEN_BG,GREEN)}
    if st in m: cell.fill=m[st][0]; cell.font=m[st][1]

# ┌──────────────────────────────────────────────────────────┐
# │  עיבוד הנתונים                                           │
# └──────────────────────────────────────────────────────────┘
def build_report(json_path, output_path):
    print(f"📂 קורא: {json_path}")
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # פרסור פקקים
    jams_raw = data.get('jams', [])
    recs = []
    for j in jams_raw:
        s = j.get('street','')
        if not s: continue
        pm = j.get('pubMillis',0)
        dt = datetime.fromtimestamp(pm/1000) if pm else None
        spd = j.get('speedKMH',0) or 0
        ln = j.get('length',0) or 0
        dl = j.get('delay',0) or 0
        tt = round((ln/1000)/spd*60,2) if spd>0 else None
        recs.append({
            'street':s, 'city':j.get('city',''), 'dt':dt,
            'date':dt.strftime('%Y-%m-%d') if dt else '',
            'day':['שני','שלישי','רביעי','חמישי','שישי','שבת','ראשון'][dt.weekday()] if dt else '',
            'hour':dt.hour if dt else None,
            'tbin':tbin(dt.hour) if dt else '',
            'speed':spd,'length_m':ln,'delay_s':dl,'tt_min':tt,
            'level':j.get('level',0),
            'sn':j.get('startNode',''),'en':j.get('endNode',''),
        })

    gj = lambda streets: [r for r in recs if r['street'] in streets] if streets else []

    wb = Workbook()
    ts = data.get('startTime','')

    # ═══════════════════ SHEET 1: סיכום מסלולים ═══════════════════
    ws1 = wb.active; ws1.title='סיכום מסלולים'
    ws1.sheet_properties.tabColor='1F3864'; ws1.sheet_view.rightToLeft=True
    c1 = ['אזור','מסלול','מוצא','יעד','מרחק (ק"מ)','זמן חופשי (דק\')',
          'כיוון 1','כיוון 2','פקקים כ1','פקקים כ2',
          'השהיה כ1 (דק\')','השהיה כ2 (דק\')','זמן משוער כ1 (דק\')','זמן משוער כ2 (דק\')']
    nc1 = len(c1)

    ws1.merge_cells('A1:N1')
    ws1.cell(row=1,column=1,value='ניתוח זמני נסיעה — מסלולים מרכזיים בישראל').font = Font(name='Arial',bold=True,size=15,color='1F3864')
    ws1.cell(row=1,column=1).alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:N2')
    ws1.cell(row=2,column=1,value=f'מקור: Waze | {ts}').font = Font(name='Arial',size=10,color='666666')
    ws1.cell(row=2,column=1).alignment = Alignment(horizontal='center')

    r=4; hdr(ws1,c1,r); r+=1
    rsums = []; csec=''
    for route in ROUTES:
        if route['section']!=csec:
            csec=route['section']
            sec_row(ws1,r,nc1,{'דרום':'🔽 דרום','מרכז':'🔶 מרכז','צפון':'🔼 צפון'}.get(csec,csec))
            r+=1
        j1=gj(route['streets_dir1']); j2=gj(route['streets_dir2'])
        d1=round(sum(x['delay_s'] for x in j1 if x['delay_s']>0)/60,1)
        d2=round(sum(x['delay_s'] for x in j2 if x['delay_s']>0)/60,1)
        e1=round(route['free_flow_min']+d1,1)
        e2=round(route['free_flow_min']+d2,1) if route['dir2_label'] else '—'
        vals=[route['section'],route['name'],route['from'],route['to'],
              route['distance_km'],route['free_flow_min'],
              route['dir1_label'],route['dir2_label'] or '—',
              len(j1),len(j2) if route['dir2_label'] else '—',
              d1,d2 if route['dir2_label'] else '—',e1,e2]
        for ci,v in enumerate(vals,1): ws1.cell(row=r,column=ci,value=v)
        srow(ws1,r,nc1,r%2==0)
        color_est(ws1.cell(row=r,column=13),e1,route['free_flow_min'])
        color_est(ws1.cell(row=r,column=14),e2,route['free_flow_min'])
        rsums.append({'route':route,'j1':j1,'j2':j2,'d1':d1,'d2':d2,'e1':e1,'e2':e2})
        r+=1
    autow(ws1,nc1)

    # ═══════════════════ SHEET 2: פירוט לפי מרווח זמן ═══════════════════
    ws2 = wb.create_sheet('פירוט לפי מרווח זמן')
    ws2.sheet_properties.tabColor='2E75B6'; ws2.sheet_view.rightToLeft=True
    c2=['אזור','מסלול','כיוון','מרווח זמן','מס\' פקקים','אורך פקקים (ק"מ)',
        'השהיה מצטברת (דק\')','מהירות ממוצעת (קמ"ש)','רמת פקק ממוצעת',
        'זמן חופשי (דק\')','זמן משוער (דק\')','תוספת %','סטטוס']
    nc2=len(c2); hdr(ws2,c2); r=2
    for rs in rsums:
        rt=rs['route']
        for lk,jl in [('dir1_label',rs['j1']),('dir2_label',rs['j2'])]:
            dl=rt[lk]
            if not dl or not jl: continue
            bytb=defaultdict(list)
            for j in jl:
                if j['tbin']: bytb[j['tbin']].append(j)
            for tb in TIME_BINS:
                jj=bytb.get(tb,[])
                if not jj: continue
                tl=round(sum(x['length_m'] for x in jj)/1000,1)
                td=round(sum(x['delay_s'] for x in jj if x['delay_s']>0)/60,1)
                asp=round(statistics.mean([x['speed'] for x in jj if x['speed']>0]),1) if any(x['speed']>0 for x in jj) else 0
                alv=round(statistics.mean([x['level'] for x in jj if isinstance(x['level'],(int,float))]),1)
                ff=rt['free_flow_min']; et=round(ff+td,1)
                pa=round((td/ff)*100,1) if ff>0 else 0
                st='חריג מאוד' if pa>50 else ('עמוס' if pa>25 else ('מתון' if pa>10 else 'תקין'))
                vals=[rt['section'],rt['name'],dl,tb,len(jj),tl,td,asp,alv,ff,et,f"{pa}%",st]
                for ci,v in enumerate(vals,1): ws2.cell(row=r,column=ci,value=v)
                srow(ws2,r,nc2,r%2==0); color_status(ws2.cell(row=r,column=13),st); r+=1
    ws2.auto_filter.ref=f"A1:{get_column_letter(nc2)}{r-1}"; ws2.freeze_panes='A2'; autow(ws2,nc2)

    # ═══════════════════ SHEET 3: השוואת כיוונים ═══════════════════
    ws3 = wb.create_sheet('השוואת כיוונים')
    ws3.sheet_properties.tabColor='ED7D31'; ws3.sheet_view.rightToLeft=True
    c3=['אזור','מסלול','מרחק','זמן חופשי','כיוון 1','כיוון 2',
        'השהיה כ1','השהיה כ2','זמן כ1','זמן כ2','הפרש','כיוון עמוס','יחס']
    nc3=len(c3); hdr(ws3,c3); r=2
    for rs in rsums:
        rt=rs['route']
        if not rt['dir2_label']: continue
        e1,e2=rs['e1'],rs['e2']
        if not isinstance(e2,(int,float)): continue
        diff=round(abs(e1-e2),1); busier=rt['dir1_label'] if e1>e2 else rt['dir2_label']
        ratio=round(max(e1,e2)/min(e1,e2),2) if min(e1,e2)>0 else 1
        vals=[rt['section'],rt['name'],rt['distance_km'],rt['free_flow_min'],
              rt['dir1_label'],rt['dir2_label'],rs['d1'],rs['d2'],e1,e2,diff,busier,f"{ratio}x"]
        for ci,v in enumerate(vals,1): ws3.cell(row=r,column=ci,value=v)
        srow(ws3,r,nc3,r%2==0)
        bc=9 if e1>e2 else 10; ws3.cell(row=r,column=bc).fill=RED_BG; ws3.cell(row=r,column=bc).font=RED
        r+=1
    ws3.freeze_panes='A2'; autow(ws3,nc3)

    # ═══════════════════ SHEET 4: חריגות ═══════════════════
    ws4 = wb.create_sheet('חריגות')
    ws4.sheet_properties.tabColor='C00000'; ws4.sheet_view.rightToLeft=True
    c4=['#','אזור','מסלול','כיוון','קטע','עיר','תאריך','יום','שעה','מרווח',
        'מהירות','אורך (מ\')','השהיה (דק\')','ממוצע (דק\')','חריגה %','רמת פקק','חומרה']
    nc4=len(c4); hdr(ws4,c4)
    anoms=[]
    for rs in rsums:
        rt=rs['route']
        for lk,jl in [('dir1_label',rs['j1']),('dir2_label',rs['j2'])]:
            dl=rt[lk]
            if not dl or not jl: continue
            dels=[x['delay_s'] for x in jl if x['delay_s']>0]
            if not dels: continue
            avgd=statistics.mean(dels); stdd=statistics.stdev(dels) if len(dels)>1 else 0
            thr=avgd+1.5*stdd if stdd>0 else avgd*2
            for j in jl:
                if j['delay_s']>thr or j['speed']<5 or j['level']>=4:
                    dp=round(((j['delay_s']-avgd)/avgd)*100,1) if avgd>0 else 0
                    sv='קריטי' if j['speed']<3 or dp>200 else ('גבוה' if j['level']>=4 or dp>100 else 'בינוני')
                    anoms.append({'rt':rt['name'],'sec':rt['section'],'dir':dl,
                        'seg':f"{j['sn']} → {j['en']}",'city':j['city'],'date':j['date'],'day':j['day'],
                        'hour':f"{j['hour']:02d}:00" if j['hour'] is not None else '','tb':j['tbin'],
                        'spd':j['speed'],'ln':j['length_m'],'dm':round(j['delay_s']/60,1),
                        'am':round(avgd/60,1),'dp':dp,'lv':j['level'],'sv':sv})
    anoms.sort(key=lambda x:(-{'קריטי':3,'גבוה':2,'בינוני':1}.get(x['sv'],0),-x['dm']))
    r=2
    for i,a in enumerate(anoms,1):
        vals=[i,a['sec'],a['rt'],a['dir'],a['seg'],a['city'],a['date'],a['day'],a['hour'],a['tb'],
              a['spd'],a['ln'],a['dm'],a['am'],f"{a['dp']}%",a['lv'],a['sv']]
        for ci,v in enumerate(vals,1): ws4.cell(row=r,column=ci,value=v)
        srow(ws4,r,nc4,r%2==0)
        sc=ws4.cell(row=r,column=17)
        if a['sv']=='קריטי': sc.fill=RED_BG; sc.font=RED
        elif a['sv']=='גבוה': sc.fill=ORANGE_BG; sc.font=ORANGE
        else: sc.fill=YELLOW_BG
        r+=1
    ws4.auto_filter.ref=f"A1:{get_column_letter(nc4)}{r-1}"; ws4.freeze_panes='A2'; autow(ws4,nc4)

    # ═══════════════════ SHEET 5: פירוט פקקים ═══════════════════
    ws5 = wb.create_sheet('פירוט פקקים')
    ws5.sheet_properties.tabColor='548235'; ws5.sheet_view.rightToLeft=True
    c5=['#','אזור','מסלול','כיוון','קטע','עיר','תאריך','יום','שעה','מרווח',
        'מהירות','אורך (מ\')','השהיה (שנ\')','זמן נסיעה (דק\')','רמת פקק']
    nc5=len(c5); hdr(ws5,c5); r=2; idx=1
    for rs in rsums:
        rt=rs['route']
        for lk,jl in [('dir1_label',rs['j1']),('dir2_label',rs['j2'])]:
            dl=rt[lk]
            if not dl: continue
            for j in sorted(jl, key=lambda x: x['dt'] or datetime.min):
                vals=[idx,rt['section'],rt['name'],dl,f"{j['sn']} → {j['en']}",j['city'],
                      j['date'],j['day'],f"{j['hour']:02d}:00" if j['hour'] is not None else '',
                      j['tbin'],j['speed'],j['length_m'],j['delay_s'],j['tt_min'],j['level']]
                for ci,v in enumerate(vals,1): ws5.cell(row=r,column=ci,value=v)
                srow(ws5,r,nc5,r%2==0); r+=1; idx+=1
    ws5.auto_filter.ref=f"A1:{get_column_letter(nc5)}{r-1}"; ws5.freeze_panes='A2'; autow(ws5,nc5)

    # ═══════════════════ SHEET 6: מקרא ═══════════════════
    ws6 = wb.create_sheet('מקרא')
    ws6.sheet_properties.tabColor='666666'; ws6.sheet_view.rightToLeft=True
    legend=[['שדה','הסבר'],['זמן חופשי','זמן נסיעה תיאורטי ללא עומסים'],
            ['השהיה מצטברת','סך ההשהיות מכלל הפקקים במסלול'],
            ['זמן נסיעה משוער','זמן חופשי + השהיה מצטברת'],
            ['תוספת %','(השהיה ÷ זמן חופשי) × 100'],
            ['רמת פקק','1=זרימה, 2=מתון, 3=בינוני, 4=כבד, 5=עצירה'],
            ['',''],['סטטוס','קריטריונים'],
            ['תקין','תוספת עד 10%'],['מתון','10%–25%'],['עמוס','25%–50%'],['חריג מאוד','מעל 50%']]
    for i,rd in enumerate(legend,1):
        for j,v in enumerate(rd,1):
            c=ws6.cell(row=i,column=j,value=v); c.font=BOLD if j==1 else DATA_FONT; c.border=BORDER
            if i==1: c.font=HDR_FONT; c.fill=HDR_FILL
    autow(ws6,2)

    # שמירה
    wb.save(output_path)

    # סטטיסטיקות
    n_total = sum(len(rs['j1'])+len(rs['j2']) for rs in rsums)
    n_north = sum(len(rs['j1'])+len(rs['j2']) for rs in rsums if rs['route']['section']=='צפון')
    print(f"✅ נשמר: {output_path}")
    print(f"   מסלולים: {len(ROUTES)}")
    print(f"   פקקים משויכים: {n_total} (צפון: {n_north})")
    print(f"   חריגות: {len(anoms)}")
    print(f"   דגימה: {ts}")


# ┌──────────────────────────────────────────────────────────┐
# │  הרצה                                                    │
# └──────────────────────────────────────────────────────────┘
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    inp = sys.argv[1]
    if not os.path.exists(inp):
        print(f"❌ קובץ לא נמצא: {inp}")
        sys.exit(1)

    if len(sys.argv) >= 3:
        out = sys.argv[2]
    else:
        # שם אוטומטי לפי תאריך
        ts = datetime.now().strftime('%Y%m%d_%H%M')
        out = f"waze_report_{ts}.xlsx"

    build_report(inp, out)
